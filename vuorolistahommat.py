import xlrd
import sys
import openpyxl
import os
from openpyxl import Workbook
from openpyxl import load_workbook

# Tällä hetkellä ongelmakohtana on vakituisten vuorot koska soluissa esiintyy tyhjän sijasta vp* 
"""
Ohjelman tarkoitus on siis hakea kaikki henkilön ARTTU työpäivät työvuorolistasta, joka on 
tallennettuna samassa muodossa kuin seuraavat: Toukokuu-converted.xlsx tai Kesakuu-converted.xlsx.
Ohjelmalle annetaan komentoriviparametrina kuukausi, jonka tyovuorot halutaan hakea ja lisätä. 
Ohjelmaa ajetaan syöttämällä komentoriville esimerkiksi:
python3 vuorolistahommat.py kesakuu
Ohjelma vaatii toimiakseen
-Kaikki käsiteltävät tiedostot sijaitsevat samassa hakemistossa ohjelman kanssa.
-Tyovuorolista on muodossa Kuukausi-converted.xlsx
    *eli alkuperäinen pdf on muutettu esim netissä smallpdf.com/pdf-to-excel excel tiedostoksi
    *muutetussa tiedostossa on kahdella eri taulukolla tyovuorolistat (siksi alkukuu ja loppukuu käsitellään erikseen)
-Tyoaikakortti on aina samanlainen
"""

def laskeTyoaika(alkamisaika, loppumisaika):
        """
        Funktiolle annetaan attribuuteiksi alkamis- ja loppumisaika 
        merkkijonoina ja lasketaan niistä työaika.
        """
        #muistetaan muuttaa minuutit sadasosiksi
        alkamistunnit = float(alkamisaika[:len(alkamisaika)-3])
        alkamisminuutit = float(alkamisaika[len(alkamisaika)-2:]) / 60.0
        alkamisaika = alkamistunnit + alkamisminuutit
        loppumistunnit = float(loppumisaika[:len(loppumisaika)-3])
        loppumisminuutit = float(loppumisaika[len(loppumisaika)-2:]) / 60.0
        loppumisaika = loppumistunnit + loppumisminuutit
        tyoaika = loppumisaika - alkamisaika
        return round(tyoaika,2)

def laskeIltalisa(loppumisaika):
        """
        Koska yli klo 18 menevä työaika lasketaan kerrottuna tietyllä prosentilla,
        se pitää laskee erikseen ja sijoittaa eri soluun.
        Tässä siis lasketaan työaika, joka menee yli klo 18. Tyäajan loppumisaika
        Annetaan attribuuttina funktiolle.
        """
        if(float(loppumisaika[:2]) >= 18):
                ylimeneva_tunteina = float(loppumisaika[:2]) - 18
                ylimeneva_minuutteina = float(loppumisaika[len(loppumisaika)-2:])
                aika = ylimeneva_tunteina + ylimeneva_minuutteina / 60
                return aika
        else:
                return 0  

"""
Käsitellään komentoriviparametri sellaiseen muotoon, että saadaan haettua ja aktivoitua oikea tyovuorolista. Haetaan ja aktivoidaan myös tyoaikakortti.
Tyovuorolistasta käsitellään alkukuu ja loppukuu erikseen, koska ne ovat eri taulukoissa. Tyoaikakortista käsitellään vain ensimmäistä taulukkoa.
"""

annettu_parametri = sys.argv[1]
lopullinen_parametri = annettu_parametri[:1].upper() + annettu_parametri[1:].lower()
tyovuorolistan_nimi = "Tyovuorolistat/" + lopullinen_parametri + "-converted.xlsx"
tyovuorolista = xlrd.open_workbook(tyovuorolistan_nimi)
tyoaikakortti = load_workbook('Tyoaikakortit/Tyoaikakortti.xlsx')
tyoaikakortin_taulukko = tyoaikakortti["tunnit"]
alkukuu = tyovuorolista.sheet_by_index(0)
loppukuu = tyovuorolista.sheet_by_index(1)

"""
Seuraavaksi etsitään kaikki työvuorolistan päivät, jolloin on mahdollista olla töitä ja lisätään ne työaikakorttiin. 
Työaikakortista pitää etsiä myös ensimmäistä työvuorolistan päivää vastaava solu, josta pitää löytää y-koordinaatti, koska x-koordinaatti on aina 2
Kun on etsitty kyseinen solu, voidaan lisätä työvuorolistaan kaikki päivät.
"""
#etsitaan kaikki paivat ja lisataan ne tyoaikakorttiin
paivat = []
viikonpaiva1_vuorolistassa = alkukuu.cell_value(0,1)
#etsitaan y-koordinaatti työaikakortin ensimmäisestä tyopäivästa, koska x- koordinaatti on aina 2
viikonpaivaY = int
for i in range(9, 16):
        if(tyoaikakortin_taulukko.cell(row=i, column=2).value == viikonpaiva1_vuorolistassa):
                viikonpaivaY = i

#etsitaan kaikki kuun päivät
for i in range (1, alkukuu.ncols):
        paiva1 = alkukuu.cell_value(1, i)
        paivat.append(paiva1)
for i in range (1, loppukuu.ncols):
        paiva2 = loppukuu.cell_value(1, i)
        paivat.append(paiva2)
#lisataan etsityt paivat tyoaikakorttiin
indeksi = 0
for i in range (viikonpaivaY, viikonpaivaY + len(paivat)):
        taytettava_viikonpaiva = tyoaikakortin_taulukko.cell(row=i, column=3)
        taytettava_viikonpaiva.value = paivat[indeksi]
        indeksi += 1

"""
Etsitään käsiteltävän henkilön nimeä (ARTTU) vastaava solu työvuorolistasta, jotta päästään käsiksi kyseisen henkilön kuukauden työvuoroihin.
Kyseisessä etsinnässä tiedetään MYÖS, että x-koordinaatti on aina 0. Kun solu tiedetään, voidaan etsiä kaikki kyseisen henkilön työvuorot
ja lisätä ne suoraan työaikakorttiin. Etsintä pitää tehdä erikseen alkukuulle ja loppukuulle. 
"""

#otetaan kasittelyyn alkukuun tyoajat
print("Alkukuun vuorot:")
#etsitaan y-koordinaatti
arttuY1 = int
for i in range (0, alkukuu.nrows):
        koordinaatti1 = alkukuu.cell_value(i, 0)
        if(koordinaatti1 == "ARTTU"):
                arttuY1 = i

#etsitaan tyotunnit ja samalla myos lisataan suoraan tyoaikakorttiin
for i in range(1, alkukuu.ncols):
        viikonpaiva1 = alkukuu.cell_value(0, i)
        day1 = alkukuu.cell_value(1, i)
        cell1 = alkukuu.cell_value(arttuY1, i)
        alkamisaika = cell1[:len(cell1)-7]
        loppumisaika = cell1[len(cell1)-5:]
        if (cell1 != ""):
                print(viikonpaiva1 + " " + day1 + " " + cell1)
                #lisataan tyoaikakorttiin jos ajankohta tasmaa
                for j in range(viikonpaivaY, viikonpaivaY + len(paivat)):
                #käsitellään lisäys jos päivä on oikea
                        if(tyoaikakortin_taulukko.cell(row=j, column=3).value == day1):
                                tyoaikakortin_taulukko.cell(row=j, column=4).value = alkamisaika 
                                tyoaikakortin_taulukko.cell(row=j, column=6).value = loppumisaika
                                tyoaikakortin_taulukko.cell(row=j, column=8).value = laskeTyoaika(alkamisaika, loppumisaika)
                                tyoaikakortin_taulukko.cell(row=j, column=9).value = laskeIltalisa(loppumisaika)
                                #lisätään vielä sunnuntaikorotus
                                if(viikonpaiva1 == "su"):
                                        tyoaikakortin_taulukko.cell(row=j, column=14).value = laskeTyoaika(alkamisaika, loppumisaika)

#otetaan kasittelyyn loppukuun tyoajat
print("Loppukuun vuorot:")
#etsitaan y-koordinaatti
arttuY2 = int
for i in range (0, loppukuu.nrows):
        koordinaatti2 = loppukuu.cell_value(i, 0)
        if(koordinaatti2 == "ARTTU"):
                arttuY2 = i

#etsitaan tyotunnit ja samalla myös lisätään ne työaikakorttiin
for i in range (1, loppukuu.ncols):
        viikonpaiva2 = loppukuu.cell_value(0, i)
        day2 = loppukuu.cell_value(1, i)
        cell2 = loppukuu.cell_value(arttuY2, i)
        #pitaa erotella nain koska python ei tunnista viivaa, joka on alku- ja loppuajan valissa
        alkamisaika = cell2[:len(cell2)-7] 
        loppumisaika = cell2[len(cell2)-5:]
        if (cell2 != ""):
                print(viikonpaiva2 + " " + day2 + " " + cell2)
                #lisataan tyoaikakorttiin jos ajankohta tasmaa
                for j in range(viikonpaivaY, viikonpaivaY + len(paivat)):
                #käsitellään lisäys jos päivä on oikea
                        if(tyoaikakortin_taulukko.cell(row=j, column=3).value == day2):
                                tyoaikakortin_taulukko.cell(row=j, column=4).value = alkamisaika
                                tyoaikakortin_taulukko.cell(row=j, column=6).value = loppumisaika
                                tyoaikakortin_taulukko.cell(row=j, column=8).value = laskeTyoaika(alkamisaika, loppumisaika)
                                tyoaikakortin_taulukko.cell(row=j, column=9).value = laskeIltalisa(loppumisaika)
                                if(viikonpaiva2 == "su"):
                                        tyoaikakortin_taulukko.cell(row=j, column=14).value = laskeTyoaika(alkamisaika, loppumisaika)

#Lopuksi tallennetaan uudeksi tiedostoksi ja avataan tiedosto
tallennettava_tiedosto = "Tyoaikakortit/Tyoaikakortti_" + lopullinen_parametri + "_valmis.xlsx"
tyoaikakortti.save(tallennettava_tiedosto)
print("Tyoajat on lisätty tiedostoon " + tallennettava_tiedosto)
os.system("open " + tallennettava_tiedosto)
